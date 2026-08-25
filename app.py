import io
import re
import json
import time
from datetime import datetime
from urllib.parse import quote_plus

import pandas as pd
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st

# --- Page Config ---
st.set_page_config(
    page_title="Noon SKU Extractor Dashboard",
    page_icon="🛍️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- Allowed Email Whitelist ---
ALLOWED_EMAILS = {
    "mnhussein@noon.com",
    "kmubarak@noon.com",
    "kmittal@noon.com",
    "kvihanga@noon.com",
    "kamehta@noon.com",
    "chakim@noon.com",
    "pragrawal@noon.com",
}

# --- Secure Email Detection ---
# 1. Check if user email is passed automatically via Streamlit Cloud headers
user_email = getattr(st.user, "email", None) or getattr(getattr(st, "experimental_user", None), "email", None)

# 2. Secure Login Handling
if not user_email:
    st.error("🔒 Access Denied: Authentication Required")
    st.info("You must view this app via Streamlit Cloud while logged into your workspace account.")
    st.stop()

user_email = user_email.strip().lower()

if user_email not in ALLOWED_EMAILS:
    st.error("🚫 Access Denied: Unauthorized Account")
    st.warning(f"Logged in as **{user_email}**, which is not on the authorized user list.")
    st.stop()

st.sidebar.success(f"Authenticated as: **{user_email}**")

# --- ZenRows API Key Integration ---
DEFAULT_ZENROWS_API_KEY = "a937e177ab01370d56a8fd844836a5cd7ea18486"
try:
    DEFAULT_ZENROWS_API_KEY = st.secrets.get("ZENROWS_API_KEY", DEFAULT_ZENROWS_API_KEY)
except Exception:
    pass

CATEGORY_MAP = {
    "Toys & Games": "toys-and-games",
    "Health & Nutrition": "health",
    "Home & Kitchen": "home-and-kitchen",
    "Other / Custom Slug": "custom"
}

PROXY_COUNTRY_MAP = {
    "Saudi Arabia": "sa",
    "UAE": "ae"
}

DEBUG_PREVIEW_CHARS = 3000

if "scrape_history" not in st.session_state:
    st.session_state.scrape_history = []
if "debug_log" not in st.session_state:
    st.session_state.debug_log = []

# --- STATIC SIDEBAR ---
st.sidebar.header("⚙️ Search Configuration")

country = st.sidebar.selectbox("Store Country", options=["Saudi Arabia", "UAE"], index=0)
selected_category_label = st.sidebar.selectbox("Category", options=list(CATEGORY_MAP.keys()), index=0)

if CATEGORY_MAP[selected_category_label] == "custom":
    category_slug = st.sidebar.text_input("Custom Category Slug", value="electronics")
else:
    category_slug = CATEGORY_MAP[selected_category_label]

search_term = st.sidebar.text_input("Search Query", value="saudi national day balloon")

p_col1, p_col2 = st.sidebar.columns(2)
with p_col1:
    start_page = st.sidebar.number_input("Start Page", min_value=1, value=1, step=1)
with p_col2:
    end_page = st.sidebar.number_input("End Page", min_value=1, value=2, step=1)

country_code = "saudi-en" if country == "Saudi Arabia" else "uae-en"
proxy_country_code = PROXY_COUNTRY_MAP.get(country, "sa")

st.sidebar.divider()
st.sidebar.subheader("🎯 Accuracy & Debug")
use_geo_proxy = st.sidebar.checkbox(
    "Match live-site results (Premium Proxy)",
    value=False,
    help=(
        "Routes the request through a residential IP located in the selected "
        "store's country, so you get the same catalog a real shopper there "
        "would see. Without this, ZenRows uses a generic proxy and results can "
        "differ from what you see in your own browser. Uses ZenRows Premium "
        "Proxy, which costs roughly 10-25x more credits per request."
    )
)
debug_mode = st.sidebar.checkbox(
    "🔍 Debug Mode (inspect raw API response)",
    value=False,
    help="Shows exactly what ZenRows/Noon returned for each page, so you can see what's happening if results look off."
)

with st.sidebar.expander("🔑 ZenRows API Key"):
    zenrows_api_key = st.text_input(
        "API Key",
        value=DEFAULT_ZENROWS_API_KEY,
        type="password",
        label_visibility="collapsed"
    )
    st.caption("For shared or deployed apps, set this via Streamlit secrets instead of leaving it in the script.")

st.sidebar.write("")
start_btn = st.sidebar.button("🚀 Start Extraction", type="primary", use_container_width=True)

st.sidebar.divider()
st.sidebar.subheader("📜 Run History")
if st.session_state.scrape_history:
    history_df = pd.DataFrame(st.session_state.scrape_history)
    st.sidebar.dataframe(history_df, use_container_width=True, hide_index=True)
else:
    st.sidebar.info("No runs recorded yet.")

# --- MAIN DASHBOARD AREA ---
st.title("🛍️ Noon SKU Extractor Dashboard")
st.caption("Configure parameters in the sidebar and run extraction.")

progress_placeholder = st.empty()
status_placeholder = st.empty()

metrics_cols = st.columns(3)
m_page = metrics_cols[0].empty()
m_skus = metrics_cols[1].empty()
m_eta = metrics_cols[2].empty()

m_page.metric("Page Progress", f"- / -")
m_skus.metric("SKUs Collected", "0")
m_eta.metric("Estimated ETA", "00:00")

download_area = st.container()
debug_area = st.container()


# --- SKU extraction helper --------------------------------------------------
def extract_skus_from_json(node, found=None):
    if found is None:
        found = []
    SKU_KEYS = ("sku", "sku_code", "skuCode", "product_sku", "productSku")

    if isinstance(node, dict):
        for key, value in node.items():
            if key in SKU_KEYS and isinstance(value, str) and value.strip():
                found.append(value.strip())
            elif isinstance(value, (dict, list)):
                extract_skus_from_json(value, found)
    elif isinstance(node, list):
        for item in node:
            if isinstance(item, (dict, list)):
                extract_skus_from_json(item, found)

    return found


# --- ZenRows API Engine (Internal JSON API Bypass) ---
def run_proxy_scraper(country_path, cat_slug, query, start_p, end_p, api_key, use_geo_proxy, proxy_country):
    sku_to_page = {}
    total_pages = (end_p - start_p) + 1
    start_time = time.time()
    pages_done = 0
    consecutive_empty_pages = 0

    session = requests.Session()
    encoded_query = quote_plus(query.strip())

    for current_page in range(start_p, end_p + 1):
        status_placeholder.info(f"Fetching Page {current_page} of {end_p} via Internal API...")

        target_noon_url = (
            f"https://www.noon.com/_svc/catalog/api/v3/u/{cat_slug}/"
            f"?limit=50&page={current_page}&q={encoded_query}"
        )

        params = {
            "apikey": api_key.strip(),
            "url": target_noon_url,
            "js_render": "false",
        }
        if use_geo_proxy:
            params["premium_proxy"] = "true"
            params["proxy_country"] = proxy_country

        headers = {
            "X-Zenrows-Apikey": api_key.strip()
        }

        try:
            res = session.get("https://api.zenrows.com/v1/", params=params, headers=headers, timeout=30)
            requested_url = target_noon_url

            if res.status_code != 200:
                alt_url = f"https://www.noon.com/_svc/catalog/api/v3/s/?limit=50&page={current_page}&q={encoded_query}"
                params["url"] = alt_url
                res = session.get("https://api.zenrows.com/v1/", params=params, headers=headers, timeout=30)
                requested_url = alt_url

            is_json = False
            response_preview = res.text[:DEBUG_PREVIEW_CHARS]
            items_on_page = 0
            new_items_found = 0

            if res.status_code == 200:
                try:
                    data = res.json()
                    is_json = True
                    response_preview = json.dumps(data, indent=2)[:DEBUG_PREVIEW_CHARS]
                    skus_this_page = extract_skus_from_json(data)
                    items_on_page = len(skus_this_page)
                    for sku in skus_this_page:
                        if sku not in sku_to_page:
                            sku_to_page[sku] = current_page
                            new_items_found += 1
                except Exception:
                    pass

                if items_on_page == 0:
                    found_skus = set(re.findall(r'/([A-Z0-9]{10,})/p/', res.text))
                    items_on_page = len(found_skus)
                    for sku in found_skus:
                        if sku not in sku_to_page:
                            sku_to_page[sku] = current_page
                            new_items_found += 1

            st.session_state.debug_log.append({
                "page": current_page,
                "requested_url": requested_url,
                "status_code": res.status_code,
                "response_preview": response_preview,
                "is_json": is_json,
            })

            if res.status_code == 200:
                if items_on_page == 0:
                    consecutive_empty_pages += 1
                    if len(sku_to_page) > 0:
                        status_placeholder.warning(
                            f"Page {current_page} returned no further items — reached the end of results."
                        )
                        break
                    elif consecutive_empty_pages >= 2:
                        status_placeholder.warning(
                            f"Pages {start_p}-{current_page} all returned no items. "
                            f"Turn on Debug Mode to inspect the raw response — the query, "
                            f"category, or endpoint may not be matching."
                        )
                        break
                else:
                    consecutive_empty_pages = 0
                    if new_items_found == 0:
                        status_placeholder.warning(
                            f"Page {current_page} returned {items_on_page} items, but all were "
                            f"already collected. Pagination may not be advancing — check Debug Mode."
                        )
                        break
            else:
                status_placeholder.error(f"Page {current_page} returned HTTP {res.status_code}: {res.text[:200]}")
                break

        except Exception as e:
            status_placeholder.error(f"Error on page {current_page}: {e}")
            break

        pages_done += 1
        pct = pages_done / total_pages
        progress_placeholder.progress(pct)

        elapsed = time.time() - start_time
        avg_time = elapsed / pages_done
        eta_seconds = int((total_pages - pages_done) * avg_time)
        eta_str = time.strftime("%M:%S", time.gmtime(eta_seconds)) if eta_seconds > 0 else "00:00"

        m_page.metric("Page Progress", f"{current_page} / {end_p}")
        m_skus.metric("SKUs Collected", len(sku_to_page))
        m_eta.metric("Estimated ETA", eta_str)

        time.sleep(0.5)

    return sku_to_page


# --- Excel Generator ---
def generate_excel_export(sku_to_page_map, country_path):
    wb = openpyxl.Workbook()
    ws_detailed = wb.active
    ws_detailed.title = "SKU Catalog"
    ws_summary = wb.create_sheet(title="Summary")

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    border_thin = Border(left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                         top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3'))
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    headers = ["Item #", "Product SKU", "Source Page", "Direct Link"]
    ws_detailed.append(headers)
    ws_detailed.row_dimensions[1].height = 24

    for c in range(1, 5):
        cell = ws_detailed.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, align_center, border_thin

    sorted_skus = sorted(sku_to_page_map.items(), key=lambda x: (x[1], x[0]))

    for idx, (sku, p_num) in enumerate(sorted_skus, start=1):
        row_idx = idx + 1
        product_url = f"https://www.noon.com/{country_path}/{sku}/p/"
        ws_detailed.append([idx, sku, f"Page {p_num}", product_url])
        ws_detailed.row_dimensions[row_idx].height = 20

        c1, c2, c3, c4 = [ws_detailed.cell(row=row_idx, column=i) for i in range(1, 5)]
        c1.alignment = c2.alignment = c3.alignment = align_center
        c4.alignment = align_left
        c1.border = c2.border = c3.border = c4.border = border_thin

        if row_idx % 2 == 0:
            zebra = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
            c1.fill = c2.fill = c3.fill = c4.fill = zebra

    for col in ws_detailed.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_detailed.column_dimensions[col_letter].width = max(max_len + 4, 15)

    ws_summary.append(["Page Number", "Products Count"])
    ws_summary.row_dimensions[1].height = 24
    for c in (1, 2):
        cell = ws_summary.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, align_center, border_thin

    page_counts = {}
    for _, p_num in sorted_skus:
        page_counts[p_num] = page_counts.get(p_num, 0) + 1

    s_row = 2
    for p_num in sorted(page_counts.keys()):
        ws_summary.append([f"Page {p_num}", page_counts[p_num]])
        ws_summary.row_dimensions[s_row].height = 20
        for c in (1, 2):
            cell = ws_summary.cell(row=s_row, column=c)
            cell.border, cell.alignment = border_thin, align_center
            if s_row % 2 == 0:
                cell.fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
        s_row += 1

    ws_summary.append(["Total Unique SKUs", len(sku_to_page_map)])
    ws_summary.row_dimensions[s_row].height = 22
    for c in (1, 2):
        cell = ws_summary.cell(row=s_row, column=c)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.border, cell.alignment = border_thin, align_center

    for col in ws_summary.columns:
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = 20

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


# --- Execution Trigger ---
if start_btn:
    if start_page > end_page:
        st.error("Start Page cannot be greater than End Page.")
    elif not search_term.strip():
        st.warning("Please enter a search term.")
    elif not zenrows_api_key.strip():
        st.warning("Please enter your ZenRows API key in the sidebar.")
    else:
        st_start_time = time.time()
        st.session_state.debug_log = []

        extracted_skus = run_proxy_scraper(
            country_code, category_slug, search_term, start_page, end_page,
            zenrows_api_key, use_geo_proxy, proxy_country_code
        )

        elapsed_sec = round(time.time() - st_start_time, 2)

        if not extracted_skus:
            st.error("No SKUs found. Check your search query or page range.")
        else:
            st.balloons()
            status_placeholder.success(f"Successfully extracted {len(extracted_skus)} products in {elapsed_sec} seconds!")

            st.session_state.scrape_history.append({
                "Time": datetime.now().strftime("%H:%M:%S"),
                "Store": country,
                "Query": search_term,
                "Pages": f"{start_page}-{end_page}",
                "SKUs": len(extracted_skus)
            })

            sorted_items = sorted(extracted_skus.items(), key=lambda x: (x[1], x[0]))
            csv_rows = []
            for idx, (sku, p_num) in enumerate(sorted_items, start=1):
                csv_rows.append({
                    "Item #": idx,
                    "Product SKU": sku,
                    "Source Page": f"Page {p_num}",
                    "Direct Link": f"https://www.noon.com/{country_code}/{sku}/p/"
                })
            df_csv = pd.DataFrame(csv_rows)

            excel_bytes = generate_excel_export(extracted_skus, country_code)
            fn_base = f"noon_{country_code}_{category_slug}_p{start_page}_to_p{end_page}_{search_term.replace(' ', '_')}"

            with download_area:
                st.divider()
                st.subheader("📬 Export Downloads")
                col_exp1, col_exp2 = st.columns(2)

                with col_exp1:
                    st.download_button(
                        label="📊 Download Excel (.xlsx)",
                        data=excel_bytes,
                        file_name=f"{fn_base}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )

                csv_data = df_csv.to_csv(index=False).encode('utf-8')
                with col_exp2:
                    st.download_button(
                        label="📄 Download CSV (.csv)",
                        data=csv_data,
                        file_name=f"{fn_base}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )

        if debug_mode and st.session_state.debug_log:
            with debug_area:
                st.divider()
                with st.expander("🔍 Debug: Raw API Responses", expanded=not extracted_skus):
                    for entry in st.session_state.debug_log:
                        st.markdown(f"**Page {entry['page']}** — HTTP `{entry['status_code']}`")
                        st.caption(entry['requested_url'])
                        st.code(entry['response_preview'], language="json" if entry['is_json'] else "text")
